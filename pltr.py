from openpyxl import load_workbook
from matplotlib import pyplot as plt
# %matplotlib inline
# for jupyter use

cargo_data = load_workbook('CargoData.xlsx')
center_truck_data = load_workbook('centerTruckData.xlsx')
halfway_truck_data=load_workbook('halfwayTruckData.xlsx')
ex_data = load_workbook('exData.xlsx')
im_data = load_workbook('imData.xlsx')


cargo_data_s1 = cargo_data['Sheet1']
center_truck_data_s1 = center_truck_data['Sheet1']
halfway_truck_data_s1 = halfway_truck_data['Sheet1']
ex_data_s1 = ex_data['Sheet1']
im_data_s1 = im_data['Sheet1']

cargo_time = cargo_data_s1['A'][1::80]
cargo = cargo_data_s1['B'][1::80]
cargo_array = []
cargo_time_array = []
for cell in cargo_time:
    cargo_time_array.append(cell.value)
for cell in cargo:
    cargo_array.append(cell.value)

center_truck_time = center_truck_data_s1['A'][1::125]
center_truck = center_truck_data_s1['B'][1::125]
center_truck_array = []
center_truck_time_array = []
for cell in center_truck_time:
    center_truck_time_array.append(cell.value)
for cell in center_truck:
    center_truck_array.append(cell.value)

halfway_truck_time = halfway_truck_data_s1['A'][1::250]
halfway_truck = halfway_truck_data_s1['B'][1::250]
halfway_truck_array = []
halfway_truck_time_array = []
for cell in halfway_truck_time:
    halfway_truck_time_array.append(cell.value)
for cell in halfway_truck:
    halfway_truck_array.append(cell.value)

ex_time = ex_data_s1['A'][1::80]
ex = ex_data_s1['B'][1::80]
ex_time_array = []
ex_array = []
for cell in ex_time:
    ex_time_array.append(cell.value)
for cell in ex:
    ex_array.append(cell.value)

im_time = im_data_s1['A'][1::80]
im = im_data_s1['B'][1::80]
im_time_array = []
im_array = []
for cell in im_time:
    im_time_array.append(cell.value)
for cell in im:
    im_array.append(cell.value)



plt.title("cargo-time")
plt.xlabel("time")
plt.ylabel("cargo")
plt.plot(cargo_time_array, cargo_array)
plt.show()

plt.title("center_truck-time")
plt.xlabel("time")
plt.ylabel("center_truck")
plt.plot(center_truck_time_array, center_truck_array)
plt.show()

plt.title("halfway_truck-time")
plt.xlabel("time")
plt.ylabel("halfway_truck")
plt.plot(halfway_truck_time_array, halfway_truck_array)
plt.show()

plt.title("ex-time")
plt.xlabel("time")
plt.ylabel("ex")
plt.plot(ex_time_array, ex_array)
plt.show()

plt.title("im-time")
plt.xlabel("time")
plt.ylabel("im")
plt.plot(im_time_array, im_array)
plt.show()

cargo_data.close()
center_truck_data.close()
ex_data.close()
im_data.close()

dataset=[]
for i in range(len(halfway_truck_time_array)):
    dataset.append([0,0])
    dataset[i]=[halfway_truck_time_array[i],center_truck_array[i]]

import csv

headers = ['time','trucknum']

with open('truck7.csv','w',newline='')as f:
    f_csv = csv.writer(f)
    f_csv.writerow(headers)
    f_csv.writerows(dataset)